import argparse
import os
import shutil
import tempfile
from pathlib import Path
from urllib.parse import parse_qs, urlparse

import pandas as pd
from openpyxl import load_workbook

from app.core.backup import create_backup
from app.core.change_log import append_changes


MASTER_FILE = Path("master/203local_Master_Directory.xlsx")
BATCHES_DIR = Path("enrichment/missing_website_batches")


def clean(value):
    if value is None:
        return ""

    value = str(value).strip()

    if value.lower() == "nan":
        return ""

    return value


def is_generic_facebook_profile_url(value):
    """
    Match unusable generic links such as:

        https://facebook.com/profile.php
        https://www.facebook.com/profile.php/

    Do not automatically remove profile.php?id=... links.
    """
    value = clean(value)

    if not value:
        return False

    try:
        parsed = urlparse(value)
    except ValueError:
        return False

    hostname = parsed.netloc.lower().split(":")[0]

    is_facebook = (
        hostname == "facebook.com"
        or hostname == "www.facebook.com"
        or hostname.endswith(".facebook.com")
    )

    if not is_facebook:
        return False

    path = parsed.path.rstrip("/").lower()

    if path != "/profile.php":
        return False

    query = parse_qs(parsed.query)

    # Preserve profile.php?id=... because it may identify a real page.
    return not clean(query.get("id", [""])[0])


def inspect_master():
    workbook = load_workbook(
        MASTER_FILE,
        read_only=True,
        data_only=False,
    )

    if len(workbook.sheetnames) != 1:
        workbook.close()
        raise RuntimeError(
            "Master workbook must contain exactly one worksheet."
        )

    worksheet = workbook[workbook.sheetnames[0]]

    headers = {
        clean(cell.value): cell.column
        for cell in worksheet[1]
        if clean(cell.value)
    }

    for required in ["business_id", "post_title", "facebook"]:
        if required not in headers:
            workbook.close()
            raise RuntimeError(
                f"Master worksheet is missing column: {required}"
            )

    matches = []

    for row_number in range(2, worksheet.max_row + 1):
        facebook = clean(
            worksheet.cell(
                row=row_number,
                column=headers["facebook"],
            ).value
        )

        if not is_generic_facebook_profile_url(facebook):
            continue

        matches.append(
            {
                "row_number": row_number,
                "business_id": clean(
                    worksheet.cell(
                        row=row_number,
                        column=headers["business_id"],
                    ).value
                ),
                "post_title": clean(
                    worksheet.cell(
                        row=row_number,
                        column=headers["post_title"],
                    ).value
                ),
                "facebook": facebook,
            }
        )

    sheet_name = worksheet.title
    business_rows = worksheet.max_row - 1

    workbook.close()

    return sheet_name, business_rows, matches


def clean_master(sheet_name, matches):
    backup_path = create_backup()

    workbook = load_workbook(MASTER_FILE)
    worksheet = workbook[sheet_name]

    headers = {
        clean(cell.value): cell.column
        for cell in worksheet[1]
        if clean(cell.value)
    }

    changes = []

    for match in matches:
        row_number = match["row_number"]
        cell = worksheet.cell(
            row=row_number,
            column=headers["facebook"],
        )

        current_value = clean(cell.value)

        # Recheck immediately before clearing.
        if not is_generic_facebook_profile_url(current_value):
            continue

        cell.value = ""

        changes.append(
            {
                "business_id": match["business_id"],
                "business_name": match["post_title"],
                "field": "facebook",
                "old_value": current_value,
                "new_value": "",
                "source": "Generic Facebook URL Cleanup",
                "confidence": "High",
            }
        )

    temp_path = (
        Path(tempfile.gettempdir())
        / "203local_Master_Directory_facebook_cleaned.xlsx"
    )

    if temp_path.exists():
        temp_path.unlink()

    workbook.save(temp_path)
    workbook.close()

    # Verify the temporary workbook before replacing the master.
    verification = load_workbook(
        temp_path,
        read_only=True,
        data_only=False,
    )

    if verification.sheetnames != [sheet_name]:
        verification.close()
        raise RuntimeError(
            "Temporary workbook changed the worksheet structure."
        )

    verification.close()

    replacement_path = MASTER_FILE.with_suffix(
        ".replacement.xlsx"
    )

    shutil.copy2(temp_path, replacement_path)
    os.replace(replacement_path, MASTER_FILE)
    temp_path.unlink(missing_ok=True)

    logged = append_changes(changes)

    return backup_path, len(changes), logged


def clean_research_batches(apply_changes):
    results = []

    if not BATCHES_DIR.exists():
        return results

    for batch_path in sorted(BATCHES_DIR.glob("*.csv")):
        try:
            df = pd.read_csv(
                batch_path,
                dtype=str,
            ).fillna("")
        except Exception as exc:
            results.append(
                {
                    "file": str(batch_path),
                    "matches": 0,
                    "error": str(exc),
                }
            )
            continue

        if "facebook" not in df.columns:
            continue

        mask = df["facebook"].map(
            is_generic_facebook_profile_url
        )

        match_count = int(mask.sum())

        if not match_count:
            continue

        if apply_changes:
            df.loc[mask, "facebook"] = ""

            # Remove the same bad URL if it was selected during review.
            if "discovered_primary_online_presence" in df.columns:
                primary_mask = df[
                    "discovered_primary_online_presence"
                ].map(is_generic_facebook_profile_url)

                df.loc[
                    primary_mask,
                    "discovered_primary_online_presence",
                ] = ""

                df.loc[
                    primary_mask,
                    "discovered_online_presence_type",
                ] = ""

                df.loc[
                    primary_mask,
                    "discovered_website_status",
                ] = ""

                df.loc[
                    primary_mask,
                    "research_status",
                ] = "Needs Further Research"

                df.loc[
                    primary_mask,
                    "research_notes",
                ] = (
                    "Generic Facebook profile.php URL removed. "
                    "A specific official page must be verified."
                )

            df.to_csv(batch_path, index=False)

        results.append(
            {
                "file": str(batch_path),
                "matches": match_count,
                "error": "",
            }
        )

    return results


def run(apply_changes=False):
    sheet_name, business_rows, master_matches = inspect_master()
    batch_results = clean_research_batches(
        apply_changes=apply_changes
    )

    batch_match_count = sum(
        result["matches"]
        for result in batch_results
        if not result["error"]
    )

    print()
    print("=" * 72)
    print("203local Facebook Quality Cleanup")
    print("=" * 72)
    print(f"Master:              {MASTER_FILE}")
    print(f"Worksheet:           {sheet_name}")
    print(f"Business rows:       {business_rows:,}")
    print(
        f"Master matches:      {len(master_matches):,}"
    )
    print(
        f"Research CSV matches:{batch_match_count:>8,}"
    )
    print(
        f"Mode:                "
        f"{'LIVE CLEANUP' if apply_changes else 'PREVIEW'}"
    )
    print()

    if master_matches:
        print("Master matches")
        print("-" * 72)

        for match in master_matches[:50]:
            print(
                f"{match['business_id']} | "
                f"{match['post_title']} | "
                f"{match['facebook']}"
            )

        if len(master_matches) > 50:
            print(
                f"...and {len(master_matches) - 50:,} more"
            )

    if batch_results:
        print()
        print("Affected research files")
        print("-" * 72)

        for result in batch_results:
            if result["error"]:
                print(
                    f"{result['file']} | ERROR: "
                    f"{result['error']}"
                )
            else:
                print(
                    f"{result['file']} | "
                    f"{result['matches']:,} match(es)"
                )

    if not apply_changes:
        print()
        print("No changes were made.")
        print("AI credits used: $0")
        return

    if master_matches:
        backup_path, cleared, logged = clean_master(
            sheet_name,
            master_matches,
        )

        print()
        print("=" * 72)
        print("Facebook Cleanup Complete")
        print("=" * 72)
        print(f"Backup created:      {backup_path}")
        print(f"Master fields cleared: {cleared:,}")
        print(f"Changes logged:        {logged:,}")
    else:
        print()
        print("No matching master fields required cleanup.")

    print(
        f"Research fields cleared: "
        f"{batch_match_count:,}"
    )
    print("Worksheets:              1")
    print("AI credits used:         $0")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()

    parser.add_argument(
        "--apply",
        action="store_true",
        help="Apply cleanup after previewing matches.",
    )

    args = parser.parse_args()

    run(apply_changes=args.apply)
