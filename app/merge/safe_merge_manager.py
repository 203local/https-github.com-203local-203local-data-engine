from pathlib import Path
from urllib.parse import urlparse

import pandas as pd
from openpyxl import load_workbook

from app.core.backup import create_backup
from app.core.change_log import append_changes


MASTER_FILE = Path(
    "master/203local_Master_Directory.xlsx"
)

ALLOWED_FIELDS = {
    "website",
    "primary_online_presence",
    "online_presence_type",
    "website_status",
}

NON_OFFICIAL_DOMAINS = {
    "facebook.com",
    "www.facebook.com",
    "instagram.com",
    "www.instagram.com",
    "tiktok.com",
    "www.tiktok.com",
    "linktr.ee",
    "www.linktr.ee",
    "singleplatform.com",
    "mappway.com",
    "yelp.com",
    "tripadvisor.com",
    "doordash.com",
    "grubhub.com",
    "ubereats.com",
    "twupro.com",
    "restaurantguru.com",
    "allmenus.com",
    "menupix.com",
    "usarestaurants.info",
}


def clean(value):
    if value is None:
        return ""

    value = str(value).strip()

    if value.lower() == "nan":
        return ""

    return value


def valid_url(value):
    value = clean(value)

    try:
        parsed = urlparse(value)

        return (
            parsed.scheme in {"http", "https"}
            and bool(parsed.netloc)
        )
    except ValueError:
        return False


def is_social_url(value):
    try:
        hostname = (
            urlparse(clean(value))
            .netloc
            .lower()
            .split(":")[0]
        )

        return any(
            hostname == domain
            or hostname.endswith("." + domain)
            for domain in NON_OFFICIAL_DOMAINS
        )
    except ValueError:
        return False


def add_proposed_change(
    proposed,
    rejected,
    master_row,
    master_index,
    business_id,
    business_name,
    field,
    new_value,
    source,
    confidence,
):
    if field not in ALLOWED_FIELDS:
        rejected.append(
            {
                "business_id": business_id,
                "field": field,
                "reason": "field not allowed",
            }
        )
        return

    new_value = clean(new_value)

    if not new_value:
        return

    current_value = clean(
        master_row.get(field)
    )

    if current_value:
        if current_value == new_value:
            return

        rejected.append(
            {
                "business_id": business_id,
                "field": field,
                "reason": (
                    "master field already populated"
                ),
            }
        )
        return

    if field in {
        "website",
        "primary_online_presence",
    }:
        if not valid_url(new_value):
            rejected.append(
                {
                    "business_id": business_id,
                    "field": field,
                    "reason": "invalid URL",
                }
            )
            return

    if (
        field == "website"
        and is_social_url(new_value)
    ):
        rejected.append(
            {
                "business_id": business_id,
                "field": field,
                "reason": (
                    "social URL cannot be stored "
                    "as official website"
                ),
            }
        )
        return

    proposed.append(
        {
            "master_index": master_index,
            "business_id": business_id,
            "business_name": business_name,
            "field": field,
            "old_value": current_value,
            "new_value": new_value,
            "source": source,
            "confidence": confidence,
        }
    )


def build_proposed_changes(
    master_df,
    batch_df,
):
    master_ids = (
        master_df["business_id"]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    proposed = []
    rejected = []

    for _, research_row in batch_df.iterrows():
        business_id = clean(
            research_row.get("business_id")
        )

        if not business_id:
            rejected.append(
                {
                    "business_id": "",
                    "field": "",
                    "reason": "missing business_id",
                }
            )
            continue

        matches = master_df.loc[
            master_ids.eq(business_id)
        ]

        if len(matches) != 1:
            rejected.append(
                {
                    "business_id": business_id,
                    "field": "",
                    "reason": (
                        "business_id not found"
                        if len(matches) == 0
                        else "duplicate business_id"
                    ),
                }
            )
            continue

        master_index = matches.index[0]
        master_row = master_df.loc[
            master_index
        ]

        business_name = clean(
            master_row.get("post_title")
        )

        status = clean(
            research_row.get(
                "research_status"
            )
        ).casefold()

        source = clean(
            research_row.get(
                "website_source"
            )
        ) or "Digital Presence Review"

        confidence = clean(
            research_row.get(
                "website_confidence"
            )
        ) or "High"

        completed_statuses = {
            "approved",
            "ready for merge",
            "verified",
            "social only",
            "facebook primary",
            "instagram primary",
            "no website found",
            "no online presence",
            "none found",
        }

        if status in completed_statuses:
            for social_field in (
                "facebook",
                "instagram",
            ):
                social_url = clean(
                    research_row.get(social_field)
                )

                if social_url:
                    add_proposed_change(
                        proposed,
                        rejected,
                        master_row,
                        master_index,
                        business_id,
                        business_name,
                        social_field,
                        social_url,
                        source,
                        confidence,
                    )

        if status in {
            "approved",
            "ready for merge",
            "verified",
        }:
            website = clean(
                research_row.get(
                    "discovered_website"
                )
            )

            primary_presence = clean(
                research_row.get(
                    "discovered_primary_online_presence"
                )
            ) or website

            presence_type = clean(
                research_row.get(
                    "discovered_online_presence_type"
                )
            ) or "Website"

            website_status = clean(
                research_row.get(
                    "discovered_website_status"
                )
            ) or "Verified Official Website"

            add_proposed_change(
                proposed,
                rejected,
                master_row,
                master_index,
                business_id,
                business_name,
                "website",
                website,
                source,
                confidence,
            )

            add_proposed_change(
                proposed,
                rejected,
                master_row,
                master_index,
                business_id,
                business_name,
                "primary_online_presence",
                primary_presence,
                source,
                confidence,
            )

            add_proposed_change(
                proposed,
                rejected,
                master_row,
                master_index,
                business_id,
                business_name,
                "online_presence_type",
                presence_type,
                source,
                confidence,
            )

            add_proposed_change(
                proposed,
                rejected,
                master_row,
                master_index,
                business_id,
                business_name,
                "website_status",
                website_status,
                source,
                confidence,
            )

        elif status in {
            "social only",
            "facebook primary",
            "instagram primary",
            "no website found",
        }:
            instagram = clean(
                research_row.get("instagram")
            )

            facebook = clean(
                research_row.get("facebook")
            )

            existing_primary = clean(
                master_row.get("primary_online_presence")
            )

            existing_type = clean(
                master_row.get("online_presence_type")
            )

            if status == "facebook primary":
                primary_presence = facebook
                presence_type = "Facebook"

            elif status == "instagram primary":
                primary_presence = instagram
                presence_type = "Instagram"

            elif (
                status in {"social only", "no website found"}
                and existing_primary
                and existing_type in {"Facebook", "Instagram"}
            ):
                primary_presence = existing_primary
                presence_type = existing_type

            elif instagram:
                primary_presence = instagram
                presence_type = "Instagram"

            elif facebook:
                primary_presence = facebook
                presence_type = "Facebook"

            else:
                primary_presence = ""
                presence_type = ""

            existing_website_status = clean(
                master_row.get("website_status")
            )

            if (
                status in {"social only", "no website found"}
                and existing_website_status
            ):
                website_status = existing_website_status
            elif presence_type == "Facebook":
                website_status = "Facebook Only"
            elif presence_type == "Instagram":
                website_status = "Instagram Only"
            else:
                website_status = "Social Only"

            add_proposed_change(
                proposed,
                rejected,
                master_row,
                master_index,
                business_id,
                business_name,
                "primary_online_presence",
                primary_presence,
                "Digital Presence Review",
                "High",
            )

            add_proposed_change(
                proposed,
                rejected,
                master_row,
                master_index,
                business_id,
                business_name,
                "online_presence_type",
                presence_type,
                "Digital Presence Review",
                "High",
            )

            add_proposed_change(
                proposed,
                rejected,
                master_row,
                master_index,
                business_id,
                business_name,
                "website_status",
                website_status,
                "Digital Presence Review",
                "High",
            )

        elif status in {
            "no online presence",
            "none found",
        }:
            add_proposed_change(
                proposed,
                rejected,
                master_row,
                master_index,
                business_id,
                business_name,
                "online_presence_type",
                "No Online Presence",
                "Digital Presence Review",
                "High",
            )

            add_proposed_change(
                proposed,
                rejected,
                master_row,
                master_index,
                business_id,
                business_name,
                "website_status",
                "No Online Presence",
                "Digital Presence Review",
                "High",
            )

        elif status in {
            "needs further research",
            "research needed",
            "pending",
            "",
        }:
            continue

        else:
            rejected.append(
                {
                    "business_id": business_id,
                    "field": "research_status",
                    "reason": (
                        "unrecognized research status: "
                        f"{status}"
                    ),
                }
            )

    return proposed, rejected


def write_changes(
    master_path,
    sheet_name,
    proposed,
):
    workbook = load_workbook(master_path)
    worksheet = workbook[sheet_name]

    headers = {
        clean(cell.value): cell.column
        for cell in worksheet[1]
        if clean(cell.value)
    }

    if "business_id" not in headers:
        workbook.close()
        raise ValueError(
            "Master worksheet has no "
            "business_id column."
        )

    for field in ALLOWED_FIELDS:
        if field not in headers:
            workbook.close()
            raise ValueError(
                f"Master worksheet is "
                f"missing column: {field}"
            )

    rows_by_id = {}

    for row_number in range(
        2,
        worksheet.max_row + 1,
    ):
        business_id = clean(
            worksheet.cell(
                row=row_number,
                column=headers["business_id"],
            ).value
        )

        if business_id:
            rows_by_id[
                business_id
            ] = row_number

    for change in proposed:
        row_number = rows_by_id.get(
            change["business_id"]
        )

        if not row_number:
            workbook.close()
            raise RuntimeError(
                "Business disappeared during "
                f"merge: {change['business_id']}"
            )

        worksheet.cell(
            row=row_number,
            column=headers[
                change["field"]
            ],
        ).value = change["new_value"]

    workbook.save(master_path)
    workbook.close()


def merge_research_batch(
    batch_path,
    dry_run=True,
):
    batch_path = Path(batch_path)

    if not MASTER_FILE.exists():
        raise FileNotFoundError(
            f"Master workbook not found: "
            f"{MASTER_FILE}"
        )

    if not batch_path.exists():
        raise FileNotFoundError(
            f"Research batch not found: "
            f"{batch_path}"
        )

    workbook_info = pd.ExcelFile(
        MASTER_FILE
    )

    if len(workbook_info.sheet_names) != 1:
        raise ValueError(
            "The authoritative workbook "
            "must contain exactly one worksheet."
        )

    sheet_name = (
        workbook_info.sheet_names[0]
    )

    master_df = pd.read_excel(
        MASTER_FILE,
        sheet_name=sheet_name,
        dtype=str,
    ).fillna("")

    batch_df = pd.read_csv(
        batch_path,
        dtype=str,
    ).fillna("")

    proposed, rejected = (
        build_proposed_changes(
            master_df,
            batch_df,
        )
    )

    print()
    print("=" * 72)
    print("203local Safe Merge Manager")
    print("=" * 72)
    print(f"Master:          {MASTER_FILE}")
    print(f"Worksheet:       {sheet_name}")
    print(f"Research batch:  {batch_path}")
    print(
        f"Proposed fields: "
        f"{len(proposed):,}"
    )
    print(
        f"Rejected fields: "
        f"{len(rejected):,}"
    )
    print(
        f"Mode:            "
        f"{'DRY RUN' if dry_run else 'LIVE MERGE'}"
    )
    print()

    for change in proposed[:30]:
        print(
            f"{change['business_id']} | "
            f"{change['business_name']} | "
            f"{change['field']} -> "
            f"{change['new_value']}"
        )

    if len(proposed) > 30:
        print(
            f"...and "
            f"{len(proposed) - 30:,} more"
        )

    if rejected:
        print()
        print("Rejected examples")
        print("-" * 72)

        for item in rejected[:10]:
            print(
                f"{item['business_id']} | "
                f"{item['field']} | "
                f"{item['reason']}"
            )

    if dry_run:
        print()
        print("No changes were made.")

        return {
            "proposed": len(proposed),
            "rejected": len(rejected),
            "backup": None,
            "changes_logged": 0,
        }

    if not proposed:
        print()
        print(
            "No approved changes are "
            "available to merge."
        )

        return {
            "proposed": 0,
            "rejected": len(rejected),
            "backup": None,
            "changes_logged": 0,
        }

    backup_path = create_backup()

    write_changes(
        MASTER_FILE,
        sheet_name,
        proposed,
    )

    changes_logged = append_changes(
        proposed
    )

    verification = load_workbook(
        MASTER_FILE,
        read_only=True,
    )

    if verification.sheetnames != [
        sheet_name
    ]:
        verification.close()
        raise RuntimeError(
            "Worksheet structure changed "
            "unexpectedly."
        )

    verification.close()

    print()
    print("=" * 72)
    print("Safe Merge Complete")
    print("=" * 72)
    print(
        f"Backup created:  {backup_path}"
    )
    print(
        f"Fields updated:  "
        f"{len(proposed):,}"
    )
    print(
        f"Changes logged:  "
        f"{changes_logged:,}"
    )
    print(
        f"Master updated:  {MASTER_FILE}"
    )
    print("Worksheets:      1")
    print()

    return {
        "proposed": len(proposed),
        "rejected": len(rejected),
        "backup": str(backup_path),
        "changes_logged": changes_logged,
    }
